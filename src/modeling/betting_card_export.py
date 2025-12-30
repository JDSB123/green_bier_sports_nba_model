"""
Excel export functionality for betting cards and pick tracking.
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import List, Optional
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import pandas as pd

from src.modeling.betting_card import BettingCardPick
from src.config import settings

CST = ZoneInfo("America/Chicago")

# Output directory - configurable via environment variable or default
OUTPUT_DIR = Path(
    os.getenv(
        "BETTING_CARD_OUTPUT_DIR",
        r"C:\Users\JB\Green Bier Capital\Early Stage Sport Ventures - Documents\Daily Picks"
    )
)
TRACKER_FILE = OUTPUT_DIR / "tracker_bombay711_daily_picks.xlsx"


def export_betting_card_to_excel(
    picks: List[BettingCardPick],
    target_date: datetime.date,
    summary_text: str,
) -> Path:
    """
    Export betting card to Excel with formatted rationale.
    
    Args:
        picks: List of betting card picks
        target_date: Date of the slate
        summary_text: Summary text to include
        
    Returns:
        Path to saved Excel file
    """
    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Create Excel file path
    excel_filename = f"betting_card_{target_date.strftime('%Y%m%d')}.xlsx"
    excel_path = OUTPUT_DIR / excel_filename
    
    # Create DataFrame for picks
    picks_data = []
    for pick in picks:
        # Join rationale bullet points with newlines for Excel cell
        rationale_text = "\n".join(pick.rationale)
        
        picks_data.append({
            "Game Date": pick.game_date,
            "Matchup": pick.matchup,
            "Pick Type": pick.pick_type,
            "Recommended Pick": pick.pick,
            "Market Line": pick.market_line,
            "Market Odds": pick.market_odds,
            "Model Prediction": pick.model_prediction,
            "Model Probability": f"{pick.model_probability*100:.1f}%",
            "Edge (pts)": pick.edge,
            "Expected Value": f"{pick.expected_value*100:+.1f}%" if pick.expected_value != 0 else "N/A",
            "Confidence": pick.confidence.upper(),
            "Rationale": rationale_text,  # Multi-line cell
        })
    
    df = pd.DataFrame(picks_data)
    
    # Write to Excel with formatting
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Summary sheet
        summary_df = pd.DataFrame([{"Summary": summary_text}])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        
        # Picks sheet
        df.to_excel(writer, sheet_name="Picks", index=False)
        
        # Get workbook and worksheet for formatting
        workbook = writer.book
        picks_sheet = writer.sheets["Picks"]
        summary_sheet = writer.sheets["Summary"]
        
        # Format picks sheet
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in picks_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Column widths
        column_widths = {
            "A": 18,  # Game Date
            "B": 30,  # Matchup
            "C": 15,  # Pick Type
            "D": 25,  # Recommended Pick
            "E": 12,  # Market Line
            "F": 12,  # Market Odds
            "G": 18,  # Model Prediction
            "H": 18,  # Model Probability
            "I": 12,  # Edge
            "J": 15,  # Expected Value
            "K": 12,  # Confidence
            "L": 80,  # Rationale (wide for multi-line)
        }
        
        for col, width in column_widths.items():
            picks_sheet.column_dimensions[col].width = width
        
        # Format rationale column for multi-line text
        for row in range(2, len(df) + 2):
            rationale_cell = picks_sheet[f"L{row}"]
            rationale_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            rationale_cell.font = Font(size=10)
        
        # Format data rows
        for row in range(2, len(df) + 2):
            for col in picks_sheet[row]:
                if col.column_letter != "L":  # Rationale column already formatted
                    col.alignment = Alignment(horizontal="center", vertical="center")
        
        # Format summary sheet
        summary_sheet.column_dimensions["A"].width = 120
        summary_cell = summary_sheet["A1"]
        summary_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        summary_cell.font = Font(size=10, name="Courier New")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in picks_sheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=12):
            for cell in row:
                cell.border = thin_border
    
    print(f"\n✅ Betting card exported to: {excel_path}")
    return excel_path


def update_pick_tracker(
    picks: List[BettingCardPick],
    target_date: datetime.date,
) -> Path:
    """
    Update or create pick tracker Excel file.
    
    Adds new picks to the tracker and maintains historical record.
    
    Args:
        picks: List of betting card picks
        target_date: Date of the slate
        
    Returns:
        Path to tracker file
    """
    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    # Prepare new picks data
    new_picks_data = []
    for pick in picks:
        new_picks_data.append({
            "Date": target_date.strftime("%Y-%m-%d"),
            "Game Date": pick.game_date,
            "Matchup": pick.matchup,
            "Pick Type": pick.pick_type,
            "Recommended Pick": pick.pick,
            "Market Line": pick.market_line,
            "Market Odds": pick.market_odds,
            "Model Prediction": pick.model_prediction,
            "Model Probability": pick.model_probability,
            "Edge (pts)": pick.edge,
            "Expected Value": pick.expected_value,
            "Confidence": pick.confidence.upper(),
            "Status": "Pending",  # Will be updated next morning
            "Result": "",  # Will be filled in after game
            "Win/Loss": "",  # Will be filled in after game
            "Notes": "",  # For manual notes
        })
    
    new_df = pd.DataFrame(new_picks_data)
    
    # Load existing tracker if it exists
    if TRACKER_FILE.exists():
        try:
            existing_df = pd.read_excel(TRACKER_FILE, sheet_name="Picks")
            # Combine with new picks
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        except Exception as e:
            print(f"⚠️  Error reading existing tracker: {e}. Creating new tracker.")
            combined_df = new_df
    else:
        combined_df = new_df
    
    # Write to Excel
    with pd.ExcelWriter(TRACKER_FILE, engine='openpyxl') as writer:
        combined_df.to_excel(writer, sheet_name="Picks", index=False)
        
        # Formatting
        workbook = writer.book
        picks_sheet = writer.sheets["Picks"]
        
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Header formatting
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in picks_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Column widths
        column_widths = {
            "A": 12,  # Date
            "B": 18,  # Game Date
            "C": 30,  # Matchup
            "D": 15,  # Pick Type
            "E": 25,  # Recommended Pick
            "F": 12,  # Market Line
            "G": 12,  # Market Odds
            "H": 18,  # Model Prediction
            "I": 18,  # Model Probability
            "J": 12,  # Edge
            "K": 15,  # Expected Value
            "L": 12,  # Confidence
            "M": 12,  # Status
            "N": 12,  # Result
            "O": 12,  # Win/Loss
            "P": 30,  # Notes
        }
        
        for col, width in column_widths.items():
            picks_sheet.column_dimensions[col].width = width
        
        # Format data rows
        for row in range(2, len(combined_df) + 2):
            for col in picks_sheet[row]:
                col.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in picks_sheet.iter_rows(min_row=1, max_row=len(combined_df) + 1, min_col=1, max_col=16):
            for cell in row:
                cell.border = thin_border
    
    print(f"✅ Pick tracker updated: {TRACKER_FILE}")
    return TRACKER_FILE


def export_betting_card_to_html(
    picks: List[BettingCardPick],
    target_date: datetime.date,
    summary_text: str,
) -> Path:
    """
    Export betting card to clean HTML format for easy reading.

    Args:
        picks: List of betting card picks
        target_date: Date of the slate
        summary_text: Summary text to include

    Returns:
        Path to saved HTML file
    """
    # Ensure output directory exists
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Create HTML file path
    html_filename = f"betting_card_{target_date.strftime('%Y%m%d')}.html"
    html_path = OUTPUT_DIR / html_filename

    # Generate HTML content
    html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>NBA Betting Card - {target_date.strftime('%B %d, %Y')}</title>
    <style>
        body {{
            font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
            line-height: 1.6;
            color: #333;
            max-width: 1200px;
            margin: 0 auto;
            padding: 20px;
            background-color: #f5f5f5;
        }}

        .header {{
            text-align: center;
            background: linear-gradient(135deg, #1e3c72, #2a5298);
            color: white;
            padding: 30px;
            border-radius: 10px;
            margin-bottom: 30px;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }}

        .header h1 {{
            margin: 0;
            font-size: 2.5em;
            font-weight: 300;
        }}

        .header p {{
            margin: 10px 0 0 0;
            opacity: 0.9;
            font-size: 1.1em;
        }}

        .summary-table {{
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin-bottom: 30px;
        }}

        .summary-table table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .summary-table th {{
            background: #2c3e50;
            color: white;
            padding: 15px 10px;
            text-align: left;
            font-weight: 600;
            font-size: 0.9em;
            border-bottom: 2px solid #34495e;
        }}

        .summary-table td {{
            padding: 12px 10px;
            border-bottom: 1px solid #ecf0f1;
            font-size: 0.9em;
        }}

        .summary-table tr:nth-child(even) {{
            background-color: #f8f9fa;
        }}

        .summary-table tr:hover {{
            background-color: #e8f4f8;
        }}

        .game-time {{
            font-weight: 600;
            color: #2c3e50;
        }}

        .matchup {{
            font-weight: 500;
            color: #34495e;
        }}

        .pick-type {{
            display: inline-block;
            padding: 3px 8px;
            border-radius: 12px;
            font-size: 0.8em;
            font-weight: 600;
            text-transform: uppercase;
        }}

        .pick-type-spread {{ background: #3498db; color: white; }}
        .pick-type-total {{ background: #e74c3c; color: white; }}

        .pick-side {{
            display: inline-block;
            padding: 4px 10px;
            border-radius: 4px;
            font-size: 0.85em;
            font-weight: 700;
            text-transform: uppercase;
        }}

        .side-away {{ background: #9b59b6; color: white; }}
        .side-home {{ background: #f39c12; color: white; }}
        .side-over {{ background: #27ae60; color: white; }}
        .side-under {{ background: #e74c3c; color: white; }}

        .model-vs-market {{
            font-family: 'Courier New', monospace;
            background: #ecf0f1;
            padding: 4px 8px;
            border-radius: 3px;
            font-size: 0.8em;
            white-space: nowrap;
        }}

        .recommended-pick {{
            font-weight: 600;
            color: #2c3e50;
            font-size: 0.9em;
        }}

        .confidence {{
            font-weight: 700;
            text-align: center;
        }}

        .high-confidence {{
            background-color: #d4edda !important;
            border-left: 4px solid #28a745;
        }}

        .high-confidence .confidence {{
            color: #28a745;
        }}

        .conflict-warning {{
            background-color: #fff3cd !important;
            border-left: 4px solid #ffc107;
        }}

        .conflict-warning .recommended-pick::before {{
            content: "⚠️ ";
            color: #856404;
        }}

        .detailed-section {{
            background: white;
            border-radius: 8px;
            padding: 25px;
            margin-bottom: 20px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}

        .detailed-section h2 {{
            color: #2c3e50;
            border-bottom: 2px solid #3498db;
            padding-bottom: 10px;
            margin-top: 0;
        }}

        .pick-item {{
            border-left: 4px solid #3498db;
            padding-left: 20px;
            margin-bottom: 25px;
            background: #f8f9fa;
            padding: 20px;
            border-radius: 5px;
        }}

        .pick-header {{
            font-size: 1.2em;
            font-weight: 600;
            color: #2c3e50;
            margin-bottom: 10px;
        }}

        .pick-details {{
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(300px, 1fr));
            gap: 15px;
            margin-bottom: 15px;
        }}

        .detail-item {{
            background: white;
            padding: 10px 15px;
            border-radius: 5px;
            border: 1px solid #dee2e6;
        }}

        .detail-label {{
            font-weight: 600;
            color: #6c757d;
            font-size: 0.85em;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .detail-value {{
            font-size: 1.1em;
            color: #495057;
            margin-top: 3px;
        }}

        .rationale {{
            background: #f8f9fa;
            border-left: 4px solid #ffc107;
            padding: 15px;
            border-radius: 0 5px 5px 0;
            margin-top: 15px;
        }}

        .rationale ul {{
            margin: 0;
            padding-left: 20px;
        }}

        .rationale li {{
            margin-bottom: 8px;
            line-height: 1.5;
        }}

        .confidence-high {{
            color: #28a745;
            font-weight: 600;
        }}

        .confidence-medium {{
            color: #ffc107;
            font-weight: 600;
        }}

        .confidence-low {{
            color: #dc3545;
            font-weight: 600;
        }}

        .footer {{
            text-align: center;
            margin-top: 40px;
            padding: 20px;
            color: #6c757d;
            border-top: 1px solid #dee2e6;
        }}

        @media (max-width: 768px) {{
            .summary-table table {{
                font-size: 0.8em;
            }}

            .summary-table th,
            .summary-table td {{
                padding: 8px 5px;
            }}

            .pick-details {{
                grid-template-columns: 1fr;
            }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🏀 NBA Betting Card</h1>
        <p>{target_date.strftime('%A, %B %d, %Y')}</p>
    </div>

    <div class="summary-table">
        <table>
            <thead>
                <tr>
                    <th>Time</th>
                    <th>Game</th>
                    <th>Bet Type</th>
                    <th>Side</th>
                    <th>What To Bet</th>
                    <th>Model → Market</th>
                    <th>Conf</th>
                </tr>
            </thead>
            <tbody>"""

    # No cross-market conflict checks for spread/total-only surface.
    conflict_picks = set()
    # Add table rows
    for pick in picks:
        # Determine pick type class
        pick_type_class = "pick-type-spread"
        if "total" in pick.pick_type.lower():
            pick_type_class = "pick-type-total"

        # Determine side class
        side_class = ""
        if pick.pick_side in ["AWAY"]:
            side_class = "side-away"
        elif pick.pick_side in ["HOME"]:
            side_class = "side-home"
        elif pick.pick_side in ["OVER"]:
            side_class = "side-over"
        elif pick.pick_side in ["UNDER"]:
            side_class = "side-under"

        # Format model vs market more clearly
        if pick.pick_type in ["FG Spread", "1H Spread"]:
            model_vs_market = f"Model: {pick.model_prediction:+.1f} → Line: {pick.market_line:+.1f}"
        elif pick.pick_type in ["FG Total", "1H Total"]:
            model_vs_market = f"Model: {pick.model_prediction:.1f} → Line: {pick.market_line:.1f}"
        else:
            model_vs_market = f"Model: {pick.model_prediction:.1f} vs Line: {pick.market_line:.1f}"

        # Determine row class (conflict takes priority over high-confidence)
        row_class = ""
        is_conflict = id(pick) in conflict_picks
        if is_conflict:
            row_class = "conflict-warning"
        elif pick.model_probability >= 0.60 or pick.model_probability <= 0.40:
            row_class = "high-confidence"

        # Create clearer game display
        game_display = f"{pick.away_team}<br>@ {pick.home_team}" if pick.away_team and pick.home_team else pick.matchup

        # Confidence indicator
        conf_pct = f"{pick.model_probability*100:.0f}%"
        
        html_content += f"""
                <tr class="{row_class}">
                    <td class="game-time">{pick.game_date}</td>
                    <td class="matchup">{game_display}</td>
                    <td><span class="pick-type {pick_type_class}">{pick.pick_type}</span></td>
                    <td><span class="pick-side {side_class}">{pick.pick_side}</span></td>
                    <td class="recommended-pick">{pick.bet_description if pick.bet_description else pick.pick}</td>
                    <td class="model-vs-market">{model_vs_market}</td>
                    <td class="confidence">{conf_pct}</td>
                </tr>"""

    html_content += """
            </tbody>
        </table>
    </div>

    <div class="detailed-section">
        <h2>📋 Detailed Breakdown</h2>"""

    # Add detailed breakdown
    for i, pick in enumerate(picks, 1):
        # Determine confidence class
        confidence_class = "confidence-medium"
        if pick.model_probability >= 0.60 or pick.model_probability <= 0.40:
            confidence_class = "confidence-high"
        elif pick.expected_value <= -5.0:
            confidence_class = "confidence-low"

        # Determine side badge color
        side_badge = ""
        if pick.pick_side == "AWAY":
            side_badge = '<span style="background:#9b59b6;color:white;padding:3px 8px;border-radius:4px;font-weight:bold;">AWAY</span>'
        elif pick.pick_side == "HOME":
            side_badge = '<span style="background:#f39c12;color:white;padding:3px 8px;border-radius:4px;font-weight:bold;">HOME</span>'
        elif pick.pick_side == "OVER":
            side_badge = '<span style="background:#27ae60;color:white;padding:3px 8px;border-radius:4px;font-weight:bold;">OVER</span>'
        elif pick.pick_side == "UNDER":
            side_badge = '<span style="background:#e74c3c;color:white;padding:3px 8px;border-radius:4px;font-weight:bold;">UNDER</span>'

        # Check if this pick has a conflict
        is_conflict_pick = id(pick) in conflict_picks
        conflict_banner = ""
        if is_conflict_pick:
            conflict_banner = """
            <div style="background:#fff3cd;border-left:4px solid #ffc107;padding:12px;margin-bottom:15px;border-radius:5px;">
                <strong>⚠️ MODEL CONFLICT DETECTED:</strong> This pick conflicts with another pick for the same game. 
                See rationale below for details. Consider both picks carefully or skip this game.
            </div>"""
        
        html_content += f"""
        <div class="pick-item">
            <div class="pick-header">{i}. {side_badge} {pick.pick_type}</div>
            {conflict_banner}
            <div style="font-size:1.1em;font-weight:600;color:#2c3e50;margin:10px 0;padding:10px;background:#f8f9fa;border-radius:5px;">
                🎯 {pick.bet_description if pick.bet_description else pick.pick}
            </div>

            <div class="pick-details">
                <div class="detail-item">
                    <div class="detail-label">Game</div>
                    <div class="detail-value">
                        <strong>{pick.away_team if pick.away_team else 'Away'}</strong> (AWAY)<br>
                        @ <strong>{pick.home_team if pick.home_team else 'Home'}</strong> (HOME)
                    </div>
                </div>

                <div class="detail-item">
                    <div class="detail-label">Game Time</div>
                    <div class="detail-value">{pick.game_date}</div>
                </div>

                <div class="detail-item">
                    <div class="detail-label">Model Prediction</div>
                    <div class="detail-value" style="font-size:1.2em;font-weight:bold;">{pick.model_prediction:.1f}</div>
                </div>

                <div class="detail-item">
                    <div class="detail-label">Market Line/Odds</div>
                    <div class="detail-value" style="font-size:1.2em;">{pick.market_line:+.1f} ({pick.market_odds:+d})</div>
                </div>

                <div class="detail-item">
                    <div class="detail-label">Win Probability</div>
                    <div class="detail-value {confidence_class}" style="font-size:1.3em;">{pick.model_probability*100:.0f}%</div>
                </div>

                <div class="detail-item">
                    <div class="detail-label">Edge vs Market</div>
                    <div class="detail-value" style="font-size:1.2em;color:#27ae60;">{pick.edge:+.1f} pts</div>
                </div>
            </div>

            <div class="rationale">
                <strong>📊 Why This Pick:</strong>
                <ul>"""

        # Add rationale bullets
        for rationale_point in pick.rationale:
            html_content += f"<li>{rationale_point}</li>"

        html_content += """
                </ul>
            </div>
        </div>"""

    html_content += """
    </div>

    <div class="footer">
        <p>Generated on """ + datetime.now(CST).strftime('%B %d, %Y at %I:%M %p %Z') + """</p>
        <p><small>Expected ROI: +7.3% for high-confidence plays (≥60% or ≤40% probability)</small></p>
    </div>
</body>
</html>"""

    # Write HTML file
    with open(html_path, 'w', encoding='utf-8') as f:
        f.write(html_content)

    print(f"\n✅ HTML betting card exported to: {html_path}")
    return html_path


def update_tracker_results(
    target_date: Optional[datetime.date] = None,
) -> None:
    """
    Update tracker with results from previous day's picks.

    This should be run the following morning at 8am CST to update
    the status and results of picks from the previous day.

    Args:
        target_date: Date to update (defaults to yesterday)
    """
    if target_date is None:
        now_cst = datetime.now(CST)
        target_date = (now_cst - timedelta(days=1)).date()

    if not TRACKER_FILE.exists():
        print(f"⚠️  Tracker file not found: {TRACKER_FILE}")
        return

    try:
        df = pd.read_excel(TRACKER_FILE, sheet_name="Picks")

        # Filter to picks from target_date that are still pending
        target_date_str = target_date.strftime("%Y-%m-%d")
        pending_picks = df[
            (df["Date"] == target_date_str) &
            (df["Status"] == "Pending")
        ].copy()

        if len(pending_picks) == 0:
            print(f"✅ No pending picks found for {target_date_str}")
            return

        print(f"📊 Found {len(pending_picks)} pending picks for {target_date_str}")
        print("⚠️  Manual update required: Please update results in the tracker file.")
        print(f"   File: {TRACKER_FILE}")
        print("   Columns to update: Status, Result, Win/Loss, Notes")

        # Update status to "Completed" (results need to be filled manually)
        df.loc[
            (df["Date"] == target_date_str) & (df["Status"] == "Pending"),
            "Status"
        ] = "Completed"

        # Save updated tracker
        with pd.ExcelWriter(TRACKER_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Picks", index=False)

        print(f"✅ Updated {len(pending_picks)} picks to 'Completed' status")
        print("   Please manually fill in Result, Win/Loss, and Notes columns")

    except Exception as e:
        print(f"❌ Error updating tracker: {e}")
        import traceback
        traceback.print_exc()

